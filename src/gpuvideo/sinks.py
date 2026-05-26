"""Destinos de eventos: para onde os eventos das soluções são gravados.

Suporta CSV (abre direto no Excel), XLSX (openpyxl), JSONL e stdout. Cada câmera
pode ter o seu (`events:` no YAML). Thread-safe (uma instância por câmera).
"""
from __future__ import annotations

import json
import os
import threading
import time
from typing import Optional

_COLS = ["timestamp", "datetime", "camera", "type", "solution", "track_id", "label", "data"]


def _row(camera_id, event):
    return [round(event.t, 3), time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(event.t)),
            camera_id, event.type, event.solution, event.track_id, event.label,
            json.dumps(event.data, ensure_ascii=False) if event.data else ""]


class EventSink:
    def write(self, camera_id, event):
        ...

    def close(self):
        pass


class StdoutSink(EventSink):
    def write(self, camera_id, event):
        print(f"[{camera_id}] {event}", flush=True)


class CsvSink(EventSink):
    def __init__(self, path):
        self.path = path
        self._lock = threading.Lock()
        d = os.path.dirname(path)
        if d:
            os.makedirs(d, exist_ok=True)
        new = not os.path.exists(path) or os.path.getsize(path) == 0
        self._f = open(path, "a", newline="", encoding="utf-8")
        import csv
        self._w = csv.writer(self._f)
        if new:
            self._w.writerow(_COLS)
            self._f.flush()

    def write(self, camera_id, event):
        with self._lock:
            self._w.writerow(_row(camera_id, event))
            self._f.flush()

    def close(self):
        with self._lock:
            self._f.close()


class JsonlSink(EventSink):
    def __init__(self, path):
        self.path = path
        self._lock = threading.Lock()
        d = os.path.dirname(path)
        if d:
            os.makedirs(d, exist_ok=True)
        self._f = open(path, "a", encoding="utf-8")

    def write(self, camera_id, event):
        rec = {"timestamp": event.t, "camera": camera_id, "type": event.type,
               "solution": event.solution, "track_id": event.track_id,
               "label": event.label, "data": event.data}
        with self._lock:
            self._f.write(json.dumps(rec, ensure_ascii=False) + "\n")
            self._f.flush()

    def close(self):
        with self._lock:
            self._f.close()


class XlsxSink(EventSink):
    """Excel via openpyxl. Salva em lote (a cada N eventos) e no close()."""

    def __init__(self, path, flush_every=20):
        try:
            from openpyxl import Workbook, load_workbook  # noqa
        except ModuleNotFoundError as e:
            raise RuntimeError("XLSX requer openpyxl: pip install gpuvideo[xlsx]") from e
        self.path = path
        self.flush_every = flush_every
        self._lock = threading.Lock()
        self._n = 0
        d = os.path.dirname(path)
        if d:
            os.makedirs(d, exist_ok=True)
        from openpyxl import Workbook, load_workbook
        if os.path.exists(path):
            self._wb = load_workbook(path)
            self._ws = self._wb.active
        else:
            self._wb = Workbook()
            self._ws = self._wb.active
            self._ws.title = "eventos"
            self._ws.append(_COLS)

    def write(self, camera_id, event):
        with self._lock:
            self._ws.append(_row(camera_id, event))
            self._n += 1
            if self._n % self.flush_every == 0:
                self._wb.save(self.path)

    def close(self):
        with self._lock:
            self._wb.save(self.path)


def make_sink(spec, camera_id, base_dir=".") -> EventSink:
    """spec: None | "stdout" | {"format": csv|xlsx|jsonl, "path": "..."} -> EventSink."""
    if spec is None or spec == "stdout":
        return StdoutSink()
    if isinstance(spec, str):  # caminho direto -> deduz formato pela extensão
        spec = {"path": spec}
    fmt = spec.get("format")
    path = spec.get("path", f"events/{camera_id}.csv")
    path = path.replace("{id}", camera_id)
    if not os.path.isabs(path):
        path = os.path.join(base_dir, path)
    if not fmt:
        fmt = os.path.splitext(path)[1].lstrip(".").lower() or "csv"
    if fmt in ("csv",):
        return CsvSink(path)
    if fmt in ("xlsx", "excel"):
        return XlsxSink(path if path.endswith(".xlsx") else path + ".xlsx")
    if fmt in ("jsonl", "json"):
        return JsonlSink(path)
    raise ValueError(f"formato de sink desconhecido: {fmt!r}")
