"""Testes dos sinks de eventos e do parsing de config YAML."""
import json
import pytest
from gpuvideo.analytics import Event
from gpuvideo.sinks import CsvSink, JsonlSink, StdoutSink, make_sink
from gpuvideo.pipeline import build_solution, load_config, _truthy


def ev():
    return Event("count_in", "fluxo", 1700000000.0, track_id=7, label="person", data={"x": 1})


def test_csv_sink(tmp_path):
    p = tmp_path / "e.csv"
    s = CsvSink(str(p))
    s.write("cam1", ev())
    s.close()
    lines = p.read_text().strip().splitlines()
    assert lines[0].startswith("timestamp,")        # header
    assert "cam1" in lines[1] and "count_in" in lines[1] and "fluxo" in lines[1]


def test_jsonl_sink(tmp_path):
    p = tmp_path / "e.jsonl"
    s = JsonlSink(str(p))
    s.write("cam2", ev())
    s.close()
    rec = json.loads(p.read_text().strip())
    assert rec["camera"] == "cam2" and rec["type"] == "count_in" and rec["track_id"] == 7


def test_make_sink(tmp_path):
    assert isinstance(make_sink("stdout", "c"), StdoutSink)
    assert isinstance(make_sink(None, "c"), StdoutSink)
    s = make_sink({"format": "csv", "path": str(tmp_path / "{id}.csv")}, "camX")
    assert isinstance(s, CsvSink) and "camX.csv" in s.path
    s.close()
    s2 = make_sink({"path": str(tmp_path / "a.jsonl")}, "c")   # formato pela extensão
    assert isinstance(s2, JsonlSink)
    s2.close()


def test_make_sink_xlsx(tmp_path):
    pytest.importorskip("openpyxl")
    from gpuvideo.sinks import XlsxSink
    s = make_sink({"format": "xlsx", "path": str(tmp_path / "e.xlsx")}, "c")
    assert isinstance(s, XlsxSink)
    s.write("c", ev())
    s.close()
    import openpyxl
    ws = openpyxl.load_workbook(str(tmp_path / "e.xlsx")).active
    assert ws.max_row == 2 and ws.cell(2, 3).value == "c"


def test_truthy():
    assert _truthy(True) and _truthy("yes")
    assert not _truthy(False) and not _truthy("false") and not _truthy(0) and not _truthy(None)


def test_build_solution():
    from gpuvideo.analytics import LineCounter, DwellZone, Heatmap, IntrusionZone
    assert isinstance(build_solution({"type": "counting", "line": [[0, .5], [1, .5]]}), LineCounter)
    assert isinstance(build_solution({"type": "dwell", "zone": [[0, 0], [1, 1]], "alert_s": 5}), DwellZone)
    assert isinstance(build_solution({"type": "heatmap"}), Heatmap)
    assert isinstance(build_solution({"type": "intrusion", "zone": [[0, 0], [1, 1]]}), IntrusionZone)
    with pytest.raises(ValueError):
        build_solution({"type": "xxx"})


def test_load_config(tmp_path):
    y = tmp_path / "c.yaml"
    y.write_text(
        "defaults: {model: yolo11n.pt, classes: [person]}\n"
        "cameras:\n"
        "  - {id: a, source: rtsp://1, solution: {type: counting, line: [[0,0.5],[1,0.5]]}}\n"
        "  - {id: b, source: rtsp://2, model: yolo11s.pt}\n")
    cams = load_config(str(y))
    assert len(cams) == 2
    assert cams[0]["model"] == "yolo11n.pt" and cams[0]["classes"] == ["person"]  # default herdado
    assert cams[1]["model"] == "yolo11s.pt"                                       # override
    assert cams[0]["id"] == "a"
